# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
import os

# [1] 페이지 설정
st.set_page_config(page_title="PRB 인건비 통합 검토", layout="wide")
st.title("📊 PRB 인건비 정합성 검토")
st.markdown("""
**PRB 파일내 성명을 기준으로 사번 추출 후 해당 등급 정합성 검토**
1. **사번 검토**: 사번추가(사번업데이트) / 사번수정(사번보정) / 동명이인 / 동일인 중복
2. **등급 정합성 검토**: 사번 기준으로 현재 해당 등급 검토
""")

# --- 표준화 및 유틸리티 로직 ---
def clean_id(val):
    if val is None or pd.isna(val): return ""
    s = str(val).split('.')[0].strip()
    return s.zfill(7) if s.isdigit() else s

def normalize_grade(val):
    if val is None or pd.isna(val): return "EMPTY"
    s = str(val).strip().upper().replace(" ", "").replace("-", "")
    if "PJ(B)계약" in s: s = s.replace("PJ(B)계약", "계약B")
    if "PJ(C)계약" in s: s = s.replace("PJ(C)계약", "계약C")
    return s

def convert_to_target_format(master_grade):
    if master_grade is None or pd.isna(master_grade): return ""
    s = str(master_grade).strip()
    s = s.replace("PJ(B)-계약", "계약B").replace("PJ(C)-계약", "계약C")
    return s

# --- UI 결과 테이블 스타일링 (웹 화면에서도 배경색으로 표시) ---
def style_p1_results(df):
    def apply_style(row):
        if row['비고'] == '사번 업데이트':
            return ['background-color: #CCE5FF; font-weight: bold;' if col == '변경 사번' else '' for col in df.columns]
        elif row['비고'] == '사번 보정':
            return ['background-color: #D5E8D4; font-weight: bold;' if col == '변경 사번' else '' for col in df.columns]
        elif row['비고'] == '동명이인':
            return ['background-color: #FFCCCC; font-weight: bold;' if col == '변경 사번' else '' for col in df.columns]
        return ['' for _ in df.columns]
    return df.style.apply(apply_style, axis=1)

if 'integrated_results' not in st.session_state:
    st.session_state.integrated_results = None

# --- 파일 업로드 UI ---
st.divider()
col1, col2 = st.columns(2)
with col1:
    master_file = st.file_uploader("기준 등급 파일 (2행 시작)", type=['xlsx'])
with col2:
    target_file = st.file_uploader("대상 PRB 파일 (6행 시작)", type=['xlsx'])

if st.sidebar.button("🧹 데이터 초기화"):
    st.session_state.integrated_results = None
    st.rerun()

# --- 메인 실행 로직 ---
if master_file and target_file:
    if st.button("🚀 PRB 데이터 검토 시작", use_container_width=True):
        try:
            with st.spinner('시나리오별 배경색 적용 및 전수 검토 중...'):
                # 1. 마스터 파일 분석
                master_bytes = master_file.getvalue()
                try:
                    df_master = pd.read_excel(io.BytesIO(master_bytes), sheet_name='SC 인원현황', header=0)
                except:
                    df_master = pd.read_excel(io.BytesIO(master_bytes), header=0)
                
                master_resources = {}
                id_to_grade_map = {} 

                for _, row in df_master.iterrows():
                    m_id = clean_id(row.iloc[0])    # A열
                    name = str(row.iloc[1]).strip() # B열
                    m_grade = row.iloc[12]          # M열
                    
                    if name not in master_resources: master_resources[name] = []
                    master_resources[name].append({'id': m_id, 'grade': m_grade, 'used': False})
                    id_to_grade_map[m_id] = m_grade

                # 2. 대상 파일 로드
                target_bytes = target_file.getvalue()
                wb = load_workbook(io.BytesIO(target_bytes))
                ws = wb['A3.자사인건비'] if 'A3.자사인건비' in wb.sheetnames else wb.active
                
                # 셀 채우기 스타일 정의 (요청하신 배경색 규격)
                fill_blue = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 사번 업데이트
                fill_green = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid") # 사번 보정
                fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # 동명이인
                fill_grade_err = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # 등급 오류

                p1_updates, p2_updates = [], []
                
                # [필터링용] 대상파일 내 (성명+사번) 조합 카운트
                target_pair_counts = {}
                for r in range(6, ws.max_row + 1):
                    t_id = clean_id(ws.cell(r, 7).value)
                    t_name = str(ws.cell(r, 8).value).strip()
                    if t_name and t_name != 'None':
                        pair = (t_name, t_id)
                        target_pair_counts[pair] = target_pair_counts.get(pair, 0) + 1

                # --- 통합 루프 시작 (6행부터) ---
                for r_idx in range(6, ws.max_row + 1):
                    name_val = ws.cell(r_idx, 8).value
                    if not name_val or str(name_val).strip() == 'None': continue
                    name = str(name_val).strip()
                    
                    original_id = clean_id(ws.cell(r_idx, 7).value)
                    original_grade = ws.cell(r_idx, 9).value
                    final_id = original_id
                    final_grade_master = None
                    note = ""

                    # --- [1단계: 사번 업데이트 및 매칭] ---
                    if name in master_resources:
                        m_list = master_resources[name]
                        is_master_name_dup = len(m_list) > 1
                        
                        # 시나리오 1: 동명이인 (기준파일에 동일 성명 2명 이상)
                        if is_master_name_dup:
                            match = next((m for m in m_list if not m['used']), m_list[0])
                            final_id = match['id']
                            final_grade_master = match['grade']
                            match['used'] = True
                            note = "동명이인"
                            
                            ws.cell(r_idx, 7).value = final_id
                            ws.cell(r_idx, 7).fill = fill_red # 빨간색 배경 (글자색 유지)
                            p1_updates.append({"행번호": r_idx, "성명": name, "기존 사번": original_id if original_id else "공란", "변경 사번": final_id, "비고": note})
                        
                        # 시나리오 2: 기준파일에 1명만 있는 경우
                        else:
                            match = m_list[0]
                            final_id = match['id']
                            final_grade_master = match['grade']
                            
                            if original_id == final_id:
                                # 성명/사번 일치 + 단일행이면 결과 제외
                                if target_pair_counts.get((name, original_id), 0) > 1:
                                    note = "동일인 중복"
                                    p1_updates.append({"행번호": r_idx, "성명": name, "기존 사번": original_id, "변경 사번": final_id, "비고": note})
                            else:
                                # 사번이 없거나 다른 경우 보정
                                ws.cell(r_idx, 7).value = final_id
                                if not original_id:
                                    note = "사번 업데이트"
                                    ws.cell(r_idx, 7).fill = fill_blue # 파란색 배경
                                else:
                                    note = "사번 보정"
                                    ws.cell(r_idx, 7).fill = fill_green # 초록색 배경
                                
                                p1_updates.append({"행번호": r_idx, "성명": name, "기존 사번": original_id if original_id else "공란", "변경 사번": final_id, "비고": note})

                    # --- [2단계: 등급 검토 (전수 조사)] ---
                    if final_id in id_to_grade_map:
                        m_grade = id_to_grade_map[final_id]
                        if normalize_grade(original_grade) != normalize_grade(m_grade):
                            fixed_grade = convert_to_target_format(m_grade)
                            ws.cell(r_idx, 9).value = fixed_grade
                            ws.cell(r_idx, 9).fill = fill_grade_err # 등급 오류 배경
                            
                            p2_updates.append({
                                "행번호": r_idx, "사번": final_id, "성명": name, 
                                "기존 등급": original_grade, "변경 등급": fixed_grade
                            })

                output = io.BytesIO()
                wb.save(output)
                
                st.session_state.integrated_results = {
                    'p1_df': pd.DataFrame(p1_updates),
                    'p2_df': pd.DataFrame(p2_updates),
                    'file_content': output.getvalue(),
                    'file_name': f"{os.path.splitext(target_file.name)[0]}_PPP_Review.xlsx"
                }

        except Exception as e:
            st.error(f"⚠️ 검토 중 오류 발생: {e}")

# --- 결과 표출 섹션 ---
if st.session_state.integrated_results:
    res = st.session_state.integrated_results
    st.divider()
    
    st.success("✅ 글자색을 유지하며 시나리오별 배경색 강조가 적용되었습니다.")
    
    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader(f"📂 1단계: 사번 확인 필요 ({len(res['p1_df'])}건)")
        if not res['p1_df'].empty:
            st.dataframe(style_p1_results(res['p1_df']), use_container_width=True)
        else:
            st.info("사번 이슈가 없습니다.")

    with col_b:
        st.subheader(f"📂 2단계: 등급 수정 내역 ({len(res['p2_df'])}건)")
        if not res['p2_df'].empty:
            st.dataframe(res['p2_df'].style.apply(lambda x: ['background-color: #FFCCCC; font-weight: bold;' if col == '변경 등급' else '' for col in res['p2_df'].columns], axis=1), use_container_width=True)
        else:
            st.info("등급 오류가 없습니다.")

    st.divider()
    st.download_button(
        label="💾 최종 검토 결과 파일 저장",
        data=res['file_content'],
        file_name=res['file_name'],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )